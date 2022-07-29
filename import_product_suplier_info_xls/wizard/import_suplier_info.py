# -*- coding: utf-8 -*-

import os
import csv
import base64
import tempfile
from odoo.exceptions import UserError
from odoo import api, fields, models, _, SUPERUSER_ID
from datetime import datetime, timedelta, date
# from xlrd import open_workbook
import openpyxl
import io


class ImportPurchaseOrder(models.TransientModel):
    _name = "wizard.import.suplier.info"

    file_data = fields.Binary('Select File', required=True, )
    file_name = fields.Char('File Name')

    def import_button(self):
        if not self.csv_validator(self.file_name):
            raise UserError(_("The file must be an .xls/.xlsx extension"))
        file_path = tempfile.gettempdir() + '/file.xlsx'
        data = self.file_data

        decoded_data = base64.b64decode(self.file_data)
        xls_filelike = io.BytesIO(decoded_data)
        workbook = openpyxl.load_workbook(xls_filelike)

        worksheet = workbook.active
        first_row = []  # The row where we store the name of the column
        for col in range(worksheet.max_column):
            first_row.append(worksheet.cell(row=1, column=col + 1).value)
        # transform the workbook to a list of dictionaries
        archive_lines = []
        for row in range(1, worksheet.max_row):
            elm = {}
            for col in range(worksheet.max_column):
                elm[first_row[col]] = worksheet.cell(row=row + 1, column=col + 1).value

            archive_lines.append(elm)

        productx = self.env['product.template'].sudo()
        suplier = self.env['product.supplierinfo'].sudo()
        vendor = self._context.get('active_id', False)

        cont = 0
        for line in archive_lines:
            cont += 1
            product_code = str(line.get('product_code', False))

            product_name = str(line.get('product_name', False))

            product_price = str(line.get('product_price', False))

            min_qty = str(line.get('min_qty', 0))
            date_start = str(line.get('date_start', False))

            date_end = str(line.get('date_end', False))
            vendor_code = str(line.get('vendor_code', ""))

            lead_time = str(line.get('lead_time', ""))
            description = str(line.get('description', ""))

            if date_start and isinstance(date_start,float):
                int_date=str(date_start).split(".")[0]
                if str(int_date) :
                    date_start =  datetime.fromtimestamp(int(int_date))
            if date_end and isinstance(date_end,float):
                int_date=str(date_end).split(".")[0]
                if str(int_date) :
                    date_end =  datetime.fromtimestamp(int(int_date))

            if product_name or product_code:
                product = productx.search(
                    ['|', ('name', '=', product_name), ('default_code', '=', product_code)])
                if product:
                    # productx.write('')
                    print()
                else:
                    product = productx.create({
                        'name': product_name,
                        'default_code': product_code
                    })

                suplier_val = {
                    'name': vendor,
                    'product_tmpl_id': product.id,
                    'price': float(product_price),
                    'min_qty': float(min_qty),
                    'date_start': date_start,
                    'date_end': date_end,
                    'vendor_code': vendor_code,
                    'lead_time': lead_time,
                    'description': description
                }

                if product and suplier_val:
                    has_suplier = suplier.search([('name','=',vendor),('product_tmpl_id','=',product.id)])
                    if has_suplier:
                        has_suplier.write(suplier_val)
                    else:
                        suplier.create(suplier_val)

    @api.model
    def csv_validator(self, xml_name):
        name, extension = os.path.splitext(xml_name)
        return True if extension == '.xls' or extension == '.xlsx' else False
